from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic import TemplateView
from django.db import transaction
from django.http import JsonResponse
from django.template.loader import render_to_string
import csv
import io
import logging
from django.db.models import Count
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Publisher, Wishlist
from .forms import PublisherForm

logger = logging.getLogger(__name__)

def _is_admin_or_subadmin(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=['admin', 'subadmin']).exists()

@login_required
def start_impersonate(request, publisher_id):
    if not _is_admin_or_subadmin(request.user):
        return HttpResponseForbidden("Not allowed")
    publisher = get_object_or_404(Publisher, pk=publisher_id)
    request.session['impersonate_publisher_id'] = publisher.id
    request.session.modified = True
    messages.info(request, f'Now viewing as publisher: {publisher.name}')
    return redirect(reverse('publishers:impersonated_dashboard', args=[publisher.id]))

@login_required
def stop_impersonate(request):
    if 'impersonate_publisher_id' in request.session:
        del request.session['impersonate_publisher_id']
        request.session.modified = True
        messages.success(request, "Returned to admin panel.")
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('publishers:publisher_list')

@login_required
def impersonated_dashboard(request, publisher_id):
    if _is_admin_or_subadmin(request.user):
        imp_id = request.session.get('impersonate_publisher_id')
        if imp_id is None or int(imp_id) != int(publisher_id):
            return HttpResponseForbidden("Impersonation not started for this publisher.")
        publisher = get_object_or_404(Publisher, pk=publisher_id)
        context = {
            'publisher': publisher,
            'impersonating': True,
        }
        return render(request, 'publishers/publisher_dashboard.html', context)
    publisher = get_object_or_404(Publisher, pk=publisher_id)
    context = {
        'publisher': publisher,
        'impersonating': False,
    }
    return render(request, 'publishers/publisher_dashboard.html', context)

def publisher_wishlist_ajax(request, publisher_id):
    try:
        # Get wishlist queryset with ordering
        wishlist_qs = Wishlist.objects.filter(publisher_id=publisher_id).order_by('-id')
        
        # Pagination - 10 items per page
        page_str = request.GET.get('page', '1')
        paginator = Paginator(wishlist_qs, 5)  # 10 wishlist items per page
        
        # Validate and convert page number
        try:
            page = int(page_str)
            if page < 1:
                page = 1
            elif page > paginator.num_pages and paginator.num_pages > 0:
                page = paginator.num_pages
        except (ValueError, TypeError):
            page = 1
        
        try:
            wishlist_items = paginator.page(page)
        except PageNotAnInteger:
            wishlist_items = paginator.page(1)
        except EmptyPage:
            wishlist_items = paginator.page(paginator.num_pages if paginator.num_pages > 0 else 1)
        
        # Render template with pagination context
        html = render_to_string('publishers/wishlist_list_fragment.html', {
            'wishlist_items': wishlist_items,
            'page_obj': wishlist_items,
            'paginator': paginator
        })
        
        return JsonResponse({'html': html})
    except Exception as e:
        return JsonResponse({'html': f'<div style="color:red">Django error: {str(e)}</div>'}, status=500)

class PublisherListView(ListView):
    model = Publisher
    template_name = 'publishers/publisher_list.html'
    context_object_name = 'publishers'

    def get_queryset(self):
        return Publisher.objects.annotate(
            wishlist_count=Count('wishlists')
        )

class PublisherCreateView(CreateView):
    model = Publisher
    form_class = PublisherForm
    template_name = 'publishers/publisher_form.html'
    success_url = reverse_lazy('publishers:publisher_list')

class PublisherUpdateView(UpdateView):
    model = Publisher
    form_class = PublisherForm
    template_name = 'publishers/publisher_form.html'
    success_url = reverse_lazy('publishers:publisher_list')

class PublisherDeleteView(DeleteView):
    model = Publisher
    template_name = 'publishers/publisher_confirm_delete.html'
    success_url = reverse_lazy('publishers:publisher_list')

class PublisherDetailView(DetailView):
    model = Publisher
    template_name = 'publishers/publisher_detail.html'
    context_object_name = 'publisher'
from django.views.generic import TemplateView

class PublisherWishlistUploadView(TemplateView):
    template_name = 'publishers/wishlist_upload.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['publishers'] = Publisher.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        # Get publisher
        publisher_id = request.POST.get('publisher')
        if not publisher_id:
            messages.error(request, "Please select a publisher.")
            return self.get(request, *args, **kwargs)
        
        publisher = get_object_or_404(Publisher, id=publisher_id)
        
        try:
            wishlist_file = request.FILES.get('wishlist_file')
            
            # Handle manual form submission
            if not wishlist_file and request.POST.get('desired_campaign'):
                return self._handle_manual_submission(request, publisher)
            
            # Handle file upload
            if wishlist_file:
                return self._handle_file_upload(request, publisher, wishlist_file)
            else:
                messages.error(request, "Please provide either a file or fill manual fields.")
                
        except Exception as e:
            logger.error(f"Unexpected error during wishlist upload: {e}")
            messages.error(request, f"Error processing upload: {str(e)}")
        
        return self.get(request, *args, **kwargs)
    
    def _handle_manual_submission(self, request, publisher):
        """Handle single manual wishlist entry"""
        try:
            def safe_decimal_from_post(val):
                try:
                    s = str(val or '').strip().replace(',', '').replace('$', '')
                    return float(s) if s else None
                except (ValueError, TypeError):
                    return None
            
            with transaction.atomic():
                Wishlist.objects.create(
                    publisher=publisher,
                    desired_campaign=(request.POST.get('desired_campaign') or '').strip()[:255],
                    geo=(request.POST.get('geo') or '').strip()[:100],
                    payout=safe_decimal_from_post(request.POST.get('payout')),
                    payable_event=(request.POST.get('payable_event') or '').strip()[:100],
                    model=(request.POST.get('model') or '').strip()[:100],
                    kpi=(request.POST.get('kpi') or '').strip()[:100],
                )
            
            messages.success(request, "Wishlist offer added successfully.")
            
        except Exception as e:
            logger.error(f"Error creating manual wishlist entry: {e}")
            messages.error(request, f"Error adding wishlist entry: {str(e)}")
        
        return self.get(request, *args, **kwargs)
    
    def _handle_file_upload(self, request, publisher, wishlist_file):
        """Handle CSV/XLSX file upload"""
        filename = wishlist_file.name.lower()
        
        if filename.endswith('.csv'):
            return self._process_csv_file(request, publisher, wishlist_file)
        elif filename.endswith('.xlsx'):
            return self._process_xlsx_file(request, publisher, wishlist_file)
        else:
            messages.error(request, "Only CSV and XLSX files are supported.")
            return self.get(request, *args, **kwargs)
    
    def _process_csv_file(self, request, publisher, wishlist_file):
        """Process CSV file with robust error handling"""
        try:
            # Try different encodings
            encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
            text = None
            
            for encoding in encodings:
                try:
                    wishlist_file.seek(0)
                    text = wishlist_file.read().decode(encoding, errors='replace')
                    break
                except UnicodeDecodeError:
                    continue
            
            if text is None:
                messages.error(request, "Could not decode CSV file. Please check file encoding.")
                return self.get(request, *args, **kwargs)
            
            # Parse CSV
            stream = io.StringIO(text)
            reader = csv.DictReader(stream)
            
            # Validate headers
            if not reader.fieldnames:
                messages.error(request, "CSV file has no headers.")
                return self.get(request, *args, **kwargs)
            
            logger.info(f"CSV headers found: {reader.fieldnames}")
            
            # Helper functions
            def get_field_value(row, field_name):
                """Get field value with flexible matching"""
                # Try exact match first
                if field_name in row:
                    return (row[field_name] or '').strip()
                
                # Try case-insensitive match
                for key in row.keys():
                    if key and key.strip().lower() == field_name.lower():
                        return (row[key] or '').strip()
                
                # Try partial match for common variations
                field_lower = field_name.lower()
                for key in row.keys():
                    if key and field_lower in key.strip().lower():
                        return (row[key] or '').strip()
                
                return ''
            
            def safe_decimal(val):
                """Safely convert value to decimal"""
                try:
                    if val is None:
                        return None
                    s = str(val).strip().replace(',', '').replace('$', '')
                    return float(s) if s else None
                except (ValueError, TypeError):
                    return None
            
            def safe_string(val, max_length=255):
                """Safely convert and truncate string"""
                if val is None:
                    return ''
                result = str(val).strip()
                return result[:max_length] if len(result) > max_length else result
            
            # Process rows
            wishlist_entries = []
            created_count = 0
            error_count = 0
            row_number = 0
            
            for row in reader:
                row_number += 1
                try:
                    # Skip completely empty rows
                    if not any(row.values()) or not any(str(v or '').strip() for v in row.values()):
                        logger.warning(f"CSV row {row_number}: Empty row, skipping")
                        continue
                    
                    # Get required field
                    desired_campaign = get_field_value(row, 'Desired Campaign')
                    if not desired_campaign:
                        logger.warning(f"CSV row {row_number}: No desired campaign, skipping")
                        continue
                    
                    # Create wishlist entry object
                    wishlist_entry = Wishlist(
                        publisher=publisher,
                        desired_campaign=safe_string(desired_campaign, 255),
                        geo=safe_string(get_field_value(row, 'Geo'), 100),
                        payout=safe_decimal(get_field_value(row, 'Payout')),
                        payable_event=safe_string(get_field_value(row, 'Payable Event'), 100),
                        model=safe_string(get_field_value(row, 'Model'), 100),
                        kpi=safe_string(get_field_value(row, 'KPI'), 100)
                    )
                    
                    wishlist_entries.append(wishlist_entry)
                    logger.info(f"CSV row {row_number}: Prepared entry for {desired_campaign}")
                    
                except Exception as e:
                    logger.error(f"CSV row {row_number}: Error processing - {e}")
                    error_count += 1
                    continue
            
            # Bulk create entries
            if wishlist_entries:
                try:
                    with transaction.atomic():
                        # Create in batches to avoid memory issues
                        batch_size = 100
                        for i in range(0, len(wishlist_entries), batch_size):
                            batch = wishlist_entries[i:i + batch_size]
                            Wishlist.objects.bulk_create(batch, ignore_conflicts=False)
                            created_count += len(batch)
                            logger.info(f"Created batch {i//batch_size + 1}: {len(batch)} wishlist entries")
                    
                    success_message = f"Successfully uploaded {created_count} wishlist entries from CSV for {publisher.name}."
                    if error_count > 0:
                        success_message += f" {error_count} rows had errors and were skipped."
                    messages.success(request, success_message)
                    
                except Exception as e:
                    logger.error(f"Error during bulk create: {e}")
                    # Fallback to individual creation
                    created_count = 0
                    with transaction.atomic():
                        for entry in wishlist_entries:
                            try:
                                entry.save()
                                created_count += 1
                            except Exception as individual_error:
                                logger.error(f"Error creating individual entry: {individual_error}")
                                error_count += 1
                    
                    if created_count > 0:
                        messages.success(request, f"Created {created_count} entries. {error_count} had errors.")
                    else:
                        messages.error(request, f"Failed to create entries. {error_count} errors occurred.")
            else:
                messages.warning(request, "No valid wishlist entries found in CSV file.")
        
        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            messages.error(request, f"Error processing CSV file: {str(e)}")
        
        return self.get(request, *args, **kwargs)
    
    def _process_xlsx_file(self, request, publisher, wishlist_file):
        """Process XLSX file with robust error handling"""
        try:
            from openpyxl import load_workbook
            
            wishlist_file.seek(0)
            wb = load_workbook(wishlist_file, read_only=True, data_only=True)
            ws = wb.active
            
            if ws.max_row < 2:
                messages.error(request, "XLSX file appears to be empty or has no data rows.")
                return self.get(request, *args, **kwargs)
            
            # Get headers from first row
            headers = []
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            for cell in header_row:
                headers.append(str(cell or '').strip())
            
            if not headers:
                messages.error(request, "XLSX file has no headers.")
                return self.get(request, *args, **kwargs)
            
            logger.info(f"XLSX headers found: {headers}")
            
            # Find column indices with flexible matching
            def find_column_index(headers, target):
                target_lower = target.lower().replace(' ', '')
                # Exact match
                for i, h in enumerate(headers):
                    if h.lower().replace(' ', '') == target_lower:
                        return i
                # Partial match
                for i, h in enumerate(headers):
                    if target_lower in h.lower().replace(' ', '') or h.lower().replace(' ', '') in target_lower:
                        return i
                return -1
            
            # Map column indices
            column_mapping = {
                'campaign': find_column_index(headers, 'Desired Campaign'),
                'geo': find_column_index(headers, 'Geo'),
                'kpi': find_column_index(headers, 'KPI'),
                'payout': find_column_index(headers, 'Payout'),
                'event': find_column_index(headers, 'Payable Event'),
                'model': find_column_index(headers, 'Model'),
            }
            
            logger.info(f"XLSX column mapping: {column_mapping}")
            
            # Helper functions
            def safe_decimal(val):
                try:
                    if val is None:
                        return None
                    s = str(val).strip().replace(',', '').replace('$', '')
                    return float(s) if s else None
                except (ValueError, TypeError):
                    return None
            
            def safe_string(val, max_length=255):
                if val is None:
                    return ''
                result = str(val).strip()
                return result[:max_length] if len(result) > max_length else result
            
            def get_cell_value(row, col_idx, max_length=255):
                if col_idx >= 0 and col_idx < len(row):
                    return safe_string(row[col_idx], max_length)
                return ''
            
            # Process data rows
            wishlist_entries = []
            created_count = 0
            error_count = 0
            row_number = 1  # Starting from 1 (header is row 0)
            
            for row_values in ws.iter_rows(min_row=2, values_only=True):
                row_number += 1
                try:
                    if not row_values:
                        continue
                    
                    # Convert to list of strings
                    row = [str(cell or '').strip() for cell in row_values]
                    
                    # Skip empty rows
                    if not any(row):
                        continue
                    
                    # Get required field
                    desired_campaign = get_cell_value(row, column_mapping['campaign'], 255)
                    if not desired_campaign:
                        logger.warning(f"XLSX row {row_number}: No desired campaign, skipping")
                        continue
                    
                    # Create wishlist entry
                    wishlist_entry = Wishlist(
                        publisher=publisher,
                        desired_campaign=desired_campaign,
                        geo=get_cell_value(row, column_mapping['geo'], 100),
                        payout=safe_decimal(get_cell_value(row, column_mapping['payout'])),
                        payable_event=get_cell_value(row, column_mapping['event'], 100),
                        model=get_cell_value(row, column_mapping['model'], 100),
                        kpi=get_cell_value(row, column_mapping['kpi'], 100)
                    )
                    
                    wishlist_entries.append(wishlist_entry)
                    logger.info(f"XLSX row {row_number}: Prepared entry for {desired_campaign}")
                    
                except Exception as e:
                    logger.error(f"XLSX row {row_number}: Error processing - {e}")
                    error_count += 1
                    continue
            
            # Bulk create entries
            if wishlist_entries:
                try:
                    with transaction.atomic():
                        # Create in batches
                        batch_size = 100
                        for i in range(0, len(wishlist_entries), batch_size):
                            batch = wishlist_entries[i:i + batch_size]
                            Wishlist.objects.bulk_create(batch, ignore_conflicts=False)
                            created_count += len(batch)
                            logger.info(f"Created batch {i//batch_size + 1}: {len(batch)} wishlist entries")
                    
                    success_message = f"Successfully uploaded {created_count} wishlist entries from XLSX for {publisher.name}."
                    if error_count > 0:
                        success_message += f" {error_count} rows had errors and were skipped."
                    messages.success(request, success_message)
                    
                except Exception as e:
                    logger.error(f"Error during bulk create: {e}")
                    # Fallback to individual creation
                    created_count = 0
                    with transaction.atomic():
                        for entry in wishlist_entries:
                            try:
                                entry.save()
                                created_count += 1
                            except Exception as individual_error:
                                logger.error(f"Error creating individual entry: {individual_error}")
                                error_count += 1
                    
                    if created_count > 0:
                        messages.success(request, f"Created {created_count} entries. {error_count} had errors.")
                    else:
                        messages.error(request, f"Failed to create entries. {error_count} errors occurred.")
            else:
                messages.warning(request, "No valid wishlist entries found in XLSX file.")
        
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            messages.error(request, f"Error processing XLSX file: {str(e)}")
        
        return self.get(request, *args, **kwargs)

class PublisherPortalView(TemplateView):
    template_name = 'publishers/portal.html'

    def get(self, request, *args, **kwargs):
        from apps.invoicing.models import Invoice
        from apps.validation.models import Validation
        invoices = Invoice.objects.all()
        validations = Validation.objects.all()
        context = self.get_context_data(invoices=invoices, validations=validations)
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        from django.contrib import messages
        from apps.invoicing.models import Invoice
        # Simulate invoice upload (add logic as needed)
        if request.FILES.get('invoice_file'):
            messages.success(request, "Invoice uploaded successfully.")
        else:
            messages.error(request, "Please select a file to upload.")
        return self.get(request, *args, **kwargs)
