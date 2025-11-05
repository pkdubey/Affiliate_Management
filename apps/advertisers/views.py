from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Advertiser
from .forms import AdvertiserForm
from django.db.models import Count
from apps.offers.models import Offer
from django.views.generic import TemplateView
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import load_workbook
import csv
import io
import logging

logger = logging.getLogger(__name__)

def advertiser_offers_ajax(request, advertiser_id):
    try:
        search_q = request.GET.get('search', '').strip()
        offers_qs = Offer.objects.filter(advertiser_id=advertiser_id).order_by('-id')

        if search_q:
            offers_qs = offers_qs.filter(
                campaign_name__icontains=search_q
            )

        page_str = request.GET.get('page', '1')
        paginator = Paginator(offers_qs, 10)
        
        try:
            page = int(page_str)
            if page < 1:
                page = 1
            elif page > paginator.num_pages and paginator.num_pages > 0:
                page = paginator.num_pages
        except (ValueError, TypeError):
            page = 1
        
        try:
            offers = paginator.page(page)
        except (PageNotAnInteger, EmptyPage):
            offers = paginator.page(1)
        
        html = render_to_string('advertisers/offer_list_fragment.html', {
            'offers': offers,
            'page_obj': offers,
            'paginator': paginator
        })
        return JsonResponse({'html': html})
    
    except Exception as e:
        return JsonResponse({'html': f'<div style="color:red">Django error: {str(e)}</div>'}, status=500)

def get_advertiser_offer_count(request, advertiser_id):
    """
    Fetch the current offer count for an advertiser
    """
    try:
        advertiser = get_object_or_404(Advertiser, id=advertiser_id)
        offer_count = advertiser.offers.count()
        return JsonResponse({'count': offer_count})
    except Exception as e:
        logger.error(f"Error fetching offer count: {e}")
        return JsonResponse({'count': 0, 'error': str(e)}, status=500)

class AdvertiserListView(ListView):
    model = Advertiser
    template_name = 'advertisers/advertiser_list.html'
    context_object_name = 'advertisers'

    def get_queryset(self):
        return Advertiser.objects.annotate(offers_count=Count('offers'))

class AdvertiserCreateView(CreateView):
    model = Advertiser
    form_class = AdvertiserForm
    template_name = 'advertisers/advertiser_form.html'
    success_url = reverse_lazy('advertisers:advertiser_list')

class AdvertiserUpdateView(UpdateView):
    model = Advertiser
    form_class = AdvertiserForm
    template_name = 'advertisers/advertiser_form.html'
    success_url = reverse_lazy('advertisers:advertiser_list')

class AdvertiserDeleteView(DeleteView):
    model = Advertiser
    template_name = 'advertisers/advertiser_confirm_delete.html'
    success_url = reverse_lazy('advertisers:advertiser_list')

class AdvertiserDetailView(DetailView):
    model = Advertiser
    template_name = 'advertisers/advertiser_detail.html'
    context_object_name = 'advertiser'

class AdvertiserOfferUploadView(TemplateView):
    template_name = 'advertisers/offer_upload.html'

    def post(self, request, *args, **kwargs):
        adv_id = request.POST.get('advertiser')
        if not adv_id:
            messages.error(request, "Missing advertiser.")
            return self.get(request, *args, **kwargs)
        
        advertiser = get_object_or_404(Advertiser, id=adv_id)
        
        try:
            offer_file = request.FILES.get('offer_file')
            
            # Manual form handling
            if not offer_file and (request.POST.get('campaign_name') or request.POST.get('geo')):
                def val(name): 
                    return (request.POST.get(name) or '').strip()
                
                with transaction.atomic():
                    Offer.objects.create(
                        advertiser=advertiser,
                        campaign_name=val('campaign_name'),
                        geo=val('geo'),
                        mmp=val('mmp'),
                        payout=val('payout') or 0,
                        payable_event=val('payable_event'),
                        model=val('model'),
                        kpi=val('kpi'),
                        title=val('campaign_name') or val('title'),
                    )
                messages.success(request, "Offer added successfully.")
                return self.get(request, *args, **kwargs)
            
            # File upload handling
            if offer_file:
                filename = offer_file.name.lower()
                rows = []
                
                # Helper functions for robust parsing
                def normalize_header(name):
                    """Normalize header name for comparison"""
                    return ''.join(str(name or '').strip().lower().split())
                
                def find_column(headers, target):
                    """Find column index by header name with fuzzy matching"""
                    target_norm = normalize_header(target)
                    # Exact match first
                    for i, h in enumerate(headers):
                        if normalize_header(h) == target_norm:
                            return i
                    # Partial match as fallback
                    for i, h in enumerate(headers):
                        if target_norm in normalize_header(h) or normalize_header(h) in target_norm:
                            return i
                    return -1
                
                def safe_decimal(val):
                    """Safely convert value to decimal/float"""
                    try:
                        if val is None:
                            return 0
                        s = str(val).strip().replace(',', '').replace('$', '')
                        return float(s) if s else 0
                    except (ValueError, TypeError):
                        return 0
                
                def safe_string(val):
                    """Safely convert value to string"""
                    if val is None:
                        return ''
                    return str(val).strip()
                
                # Process CSV files
                if filename.endswith('.csv'):
                    try:
                        # Try different encodings
                        encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']
                        text = None
                        
                        for encoding in encodings:
                            try:
                                offer_file.seek(0)  # Reset file pointer
                                text = offer_file.read().decode(encoding, errors='replace')
                                break
                            except UnicodeDecodeError:
                                continue
                        
                        if text is None:
                            messages.error(request, "Could not decode CSV file. Please check file encoding.")
                            return self.get(request, *args, **kwargs)
                        
                        # Detect delimiter
                        sample = text[:1024]
                        sniffer = csv.Sniffer()
                        try:
                            delimiter = sniffer.sniff(sample).delimiter
                        except csv.Error:
                            delimiter = ','
                        
                        stream = io.StringIO(text)
                        reader = csv.reader(stream, delimiter=delimiter)
                        
                        headers = []
                        try:
                            headers = next(reader, [])
                        except StopIteration:
                            messages.error(request, "CSV file appears to be empty.")
                            return self.get(request, *args, **kwargs)
                        
                        if not headers:
                            messages.error(request, "CSV file has no headers.")
                            return self.get(request, *args, **kwargs)
                        
                        logger.info(f"CSV Headers found: {headers}")
                        
                        # Find column indices
                        column_mapping = {
                            'campaign': find_column(headers, 'Campaign Name'),
                            'geo': find_column(headers, 'Geo'),
                            'kpi': find_column(headers, 'KPI'),
                            'mmp': find_column(headers, 'MMP'),
                            'payout': find_column(headers, 'Payout'),
                            'event': find_column(headers, 'Payable Event'),
                            'model': find_column(headers, 'Model'),
                        }
                        
                        logger.info(f"Column mapping: {column_mapping}")
                        
                        # Process each row
                        row_count = 0
                        for row_num, row in enumerate(reader, start=2):  # Start from 2 since header is row 1
                            try:
                                if not row or len(row) == 0:
                                    continue
                                
                                # Skip rows that are too short to contain required data
                                max_needed_index = max([idx for idx in column_mapping.values() if idx >= 0], default=0)
                                if len(row) <= max_needed_index:
                                    logger.warning(f"Row {row_num}: Too short, skipping")
                                    continue
                                
                                # Extract values safely
                                row_data = {
                                    'Campaign Name': safe_string(row[column_mapping['campaign']] if column_mapping['campaign'] >= 0 else ''),
                                    'Geo': safe_string(row[column_mapping['geo']] if column_mapping['geo'] >= 0 else ''),
                                    'KPI': safe_string(row[column_mapping['kpi']] if column_mapping['kpi'] >= 0 else ''),
                                    'MMP': safe_string(row[column_mapping['mmp']] if column_mapping['mmp'] >= 0 else ''),
                                    'Payout': safe_decimal(row[column_mapping['payout']] if column_mapping['payout'] >= 0 else 0),
                                    'Payable Event': safe_string(row[column_mapping['event']] if column_mapping['event'] >= 0 else ''),
                                    'Model': safe_string(row[column_mapping['model']] if column_mapping['model'] >= 0 else ''),
                                }
                                
                                # Skip rows with empty campaign name
                                if not row_data['Campaign Name']:
                                    logger.warning(f"Row {row_num}: Empty campaign name, skipping")
                                    continue
                                
                                rows.append(row_data)
                                row_count += 1
                                
                            except Exception as e:
                                logger.error(f"Error processing CSV row {row_num}: {e}")
                                continue
                        
                        logger.info(f"Processed {row_count} rows from CSV")
                        
                    except Exception as e:
                        logger.error(f"Error reading CSV file: {e}")
                        messages.error(request, f"Error reading CSV file: {e}")
                        return self.get(request, *args, **kwargs)
                
                # Process XLSX files
                elif filename.endswith('.xlsx'):
                    try:
                        offer_file.seek(0)  # Reset file pointer
                        wb = load_workbook(offer_file, read_only=True, data_only=True)
                        ws = wb.active
                        
                        if ws.max_row < 2:
                            messages.error(request, "XLSX file appears to be empty or has no data rows.")
                            return self.get(request, *args, **kwargs)
                        
                        # Get headers from first row
                        headers = []
                        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                        for cell in header_row:
                            headers.append(safe_string(cell))
                        
                        if not headers:
                            messages.error(request, "XLSX file has no headers.")
                            return self.get(request, *args, **kwargs)
                        
                        logger.info(f"XLSX Headers found: {headers}")
                        
                        # Find column indices
                        column_mapping = {
                            'campaign': find_column(headers, 'Campaign Name'),
                            'geo': find_column(headers, 'Geo'),
                            'kpi': find_column(headers, 'KPI'),
                            'mmp': find_column(headers, 'MMP'),
                            'payout': find_column(headers, 'Payout'),
                            'event': find_column(headers, 'Payable Event'),
                            'model': find_column(headers, 'Model'),
                        }
                        
                        logger.info(f"Column mapping: {column_mapping}")
                        
                        # Process data rows
                        row_count = 0
                        for row_num, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            try:
                                if not row_values:
                                    continue
                                
                                # Convert to list of strings
                                row = [safe_string(cell) for cell in row_values]
                                
                                # Skip empty rows
                                if not any(row):
                                    continue
                                
                                # Skip rows that are too short
                                max_needed_index = max([idx for idx in column_mapping.values() if idx >= 0], default=0)
                                if len(row) <= max_needed_index:
                                    logger.warning(f"Row {row_num}: Too short, skipping")
                                    continue
                                
                                # Extract values safely
                                row_data = {
                                    'Campaign Name': safe_string(row[column_mapping['campaign']] if column_mapping['campaign'] >= 0 and column_mapping['campaign'] < len(row) else ''),
                                    'Geo': safe_string(row[column_mapping['geo']] if column_mapping['geo'] >= 0 and column_mapping['geo'] < len(row) else ''),
                                    'KPI': safe_string(row[column_mapping['kpi']] if column_mapping['kpi'] >= 0 and column_mapping['kpi'] < len(row) else ''),
                                    'MMP': safe_string(row[column_mapping['mmp']] if column_mapping['mmp'] >= 0 and column_mapping['mmp'] < len(row) else ''),
                                    'Payout': safe_decimal(row[column_mapping['payout']] if column_mapping['payout'] >= 0 and column_mapping['payout'] < len(row) else 0),
                                    'Payable Event': safe_string(row[column_mapping['event']] if column_mapping['event'] >= 0 and column_mapping['event'] < len(row) else ''),
                                    'Model': safe_string(row[column_mapping['model']] if column_mapping['model'] >= 0 and column_mapping['model'] < len(row) else ''),
                                }
                                
                                # Skip rows with empty campaign name
                                if not row_data['Campaign Name']:
                                    logger.warning(f"Row {row_num}: Empty campaign name, skipping")
                                    continue
                                
                                rows.append(row_data)
                                row_count += 1
                                
                            except Exception as e:
                                logger.error(f"Error processing XLSX row {row_num}: {e}")
                                continue
                        
                        logger.info(f"Processed {row_count} rows from XLSX")
                        
                    except Exception as e:
                        logger.error(f"Error reading XLSX file: {e}")
                        messages.error(request, f"Error reading XLSX file: {e}")
                        return self.get(request, *args, **kwargs)
                
                else:
                    messages.error(request, "Unsupported file type. Use CSV or XLSX.")
                    return self.get(request, *args, **kwargs)
                
                # Create offers in bulk with transaction
                if rows:
                    created_count = 0
                    error_count = 0
                    
                    # Use bulk operations for better performance
                    offers_to_create = []
                    
                    for i, row_data in enumerate(rows):
                        try:
                            offer = Offer(
                                advertiser=advertiser,
                                campaign_name=row_data['Campaign Name'][:255],  # Ensure field length limits
                                geo=row_data['Geo'][:100] if row_data['Geo'] else '',
                                mmp=row_data['MMP'][:100] if row_data['MMP'] else '',
                                payout=row_data['Payout'],
                                payable_event=row_data['Payable Event'][:100] if row_data['Payable Event'] else '',
                                model=row_data['Model'][:100] if row_data['Model'] else '',
                                kpi=row_data['KPI'][:100] if row_data['KPI'] else '',
                                title=row_data['Campaign Name'][:255],
                            )
                            offers_to_create.append(offer)
                            
                        except Exception as e:
                            logger.error(f"Error preparing offer {i+1}: {e}")
                            error_count += 1
                            continue
                    
                    # Bulk create offers
                    if offers_to_create:
                        try:
                            with transaction.atomic():
                                # Create in batches to avoid memory issues
                                batch_size = 100
                                for i in range(0, len(offers_to_create), batch_size):
                                    batch = offers_to_create[i:i + batch_size]
                                    created_offers = Offer.objects.bulk_create(batch, ignore_conflicts=False)
                                    created_count += len(batch)
                                    logger.info(f"Created batch {i//batch_size + 1}: {len(batch)} offers")
                        
                        except Exception as e:
                            logger.error(f"Error during bulk create: {e}")
                            # Fallback to individual creation
                            created_count = 0
                            with transaction.atomic():
                                for offer in offers_to_create:
                                    try:
                                        offer.save()
                                        created_count += 1
                                    except Exception as individual_error:
                                        logger.error(f"Error creating individual offer: {individual_error}")
                                        error_count += 1
                    
                    if created_count > 0:
                        success_message = f"Successfully uploaded {created_count} offers for {advertiser.company_name}."
                        if error_count > 0:
                            success_message += f" {error_count} rows had errors and were skipped."
                        messages.success(request, success_message)
                    else:
                        messages.error(request, f"No offers were created. {error_count} rows had errors.")
                        
                else:
                    messages.warning(request, "No valid data found in the uploaded file.")
                    
            else:
                messages.error(request, "Choose a file or fill manual fields.")
                
        except Exception as e:
            logger.error(f"Unexpected error during upload: {e}")
            messages.error(request, f"Error processing upload: {e}")
        
        return self.get(request, *args, **kwargs)